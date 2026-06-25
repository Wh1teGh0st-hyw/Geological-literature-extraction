"""导出引擎 — 将分类后的地球化学数据导出为 Excel/CSV。"""

import os
from typing import List, Optional, Dict, Any
from loguru import logger

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportEngine:
    """地球化学数据导出引擎。"""

    # 94列标准模板
    EXCEL_HEADERS = [
        "矿区", "岩体", "Data source", "Sample.No", "Granite type", "Rock type",
        "成岩年龄", "error", "Method", "成矿年龄", "error", "Method",
        "X", "Y", "Hight",
        "SiO2", "TiO2", "Al2O3", "Fe2O3", "FeO", "MnO", "MgO", "CaO",
        "Na2O", "K2O", "P2O5", "L.O.I", "Total",
        "Li", "Be", "Sc", "V", "Cr", "Co", "Ni", "Cu", "Zn", "Ga",
        "Rb", "Sr", "Y", "Zr", "Nb", "Cs", "Ba", "Hf", "Ta", "Pb", "Th", "U",
        "La", "Ce", "Pr", "Nd", "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb", "Lu",
        "87Sr/86Sr(i)", "87Sr/86Sr_err",
        "143Nd/144Nd(i)", "143Nd/144Nd_err", "εNd(t)", "εNd_err",
        "TDM2(Nd)", "TDM1(Nd)",
        "176Hf/177Hf(i)", "176Hf/177Hf_err", "εHf(t)", "εHf_err",
        "TDM2(Hf)", "TDM1(Hf)",
        "ASI", "Ga/Al×10⁴", "FeOt/(FeOt+MgO)", "TZr(°C)",
    ]

    # Claude主题颜色用于Excel
    HEADER_FILL = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    DATA_FONT = Font(name="Microsoft YaHei", color="1A1A2E", size=9)
    ITYPE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")     # 浅蓝
    STYPE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")     # 浅红
    ATYPE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")     # 浅绿
    MTYPE_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")     # 浅紫
    THIN_BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def export_excel(self, samples: List[Dict[str, Any]],
                     output_path: str,
                     group_by: str = None,
                     include_indices: bool = True) -> str:
        """将样本列表导出为94列Excel文件。

        Args:
            samples: 样本字典列表 (to_flat_dict() 输出)
            output_path: 输出文件路径
            group_by: 按此字段分组到不同Sheet (如 "Granite type")
            include_indices: 是否包含计算指数列
        Returns:
            输出文件路径
        """
        if not samples:
            logger.warning("没有数据可导出")
            return ""

        wb = openpyxl.Workbook()
        # 删除默认Sheet
        wb.remove(wb.active)

        if group_by:
            # 按字段分组输出
            groups = {}
            for s in samples:
                key = str(s.get(group_by, "未分组")) or "未分组"
                if key not in groups:
                    groups[key] = []
                groups[key].append(s)

            for group_name, group_samples in sorted(groups.items()):
                # Sheet名限制31字符
                sheet_name = str(group_name)[:31]
                self._write_sheet(wb, sheet_name, group_samples, include_indices)
        else:
            self._write_sheet(wb, "Geochem Data", samples, include_indices)

        wb.save(output_path)
        logger.info(f"Excel导出完成: {output_path} ({len(samples)} 样本)")

        return output_path

    def _write_sheet(self, wb, sheet_name: str, samples: List[Dict], include_indices: bool):
        """写入单个Sheet。"""
        ws = wb.create_sheet(title=sheet_name)

        # 确定列标题
        headers = self.EXCEL_HEADERS.copy()
        if not include_indices:
            # 去掉计算指数列
            headers = headers[:-4]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

        # 写入数据
        for row_idx, sample in enumerate(samples, 2):
            for col_idx, header in enumerate(headers, 1):
                value = sample.get(header, "")
                if value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER

                # 数字对齐
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

            # 按花岗岩类型着色
            granite_type = str(sample.get("Granite type", "")).strip().upper()
            if "A-" in granite_type or "ATYPE" in granite_type:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.ATYPE_FILL
            elif "S-" in granite_type or "STYPE" in granite_type:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.STYPE_FILL
            elif "M-" in granite_type or "MTYPE" in granite_type:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.MTYPE_FILL
            elif "I-" in granite_type or "ITYPE" in granite_type:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.ITYPE_FILL

        # 冻结首行
        ws.freeze_panes = "A2"

        # 列宽自适应
        for col_idx in range(1, min(len(headers) + 1, 30)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

        # 关键列稍宽
        ws.column_dimensions['A'].width = 10   # 矿区
        ws.column_dimensions['B'].width = 16   # 岩体
        ws.column_dimensions['C'].width = 14   # Data source
        ws.column_dimensions['D'].width = 12   # Sample.No
        ws.column_dimensions['E'].width = 14   # Granite type
        ws.column_dimensions['F'].width = 16   # Rock type

    def export_csv(self, samples: List[Dict[str, Any]],
                   output_path: str,
                   include_indices: bool = True) -> str:
        """将样本列表导出为UTF-8 BOM CSV（兼容Excel中文）。"""
        if not samples:
            logger.warning("没有数据可导出")
            return ""

        headers = self.EXCEL_HEADERS.copy()
        if not include_indices:
            headers = headers[:-4]

        rows = []
        for sample in samples:
            row = [sample.get(h, "") for h in headers]
            rows.append(row)

        df = pd.DataFrame(rows, columns=headers)

        # UTF-8 BOM for Excel compatibility
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            df.to_csv(f, index=False)

        logger.info(f"CSV导出完成: {output_path} ({len(samples)} 样本)")
        return output_path

    def export_diagram_data(self, samples: List[Dict[str, Any]],
                            output_dir: str) -> Dict[str, str]:
        """导出图解数据文件。

        Returns:
            {文件类型: 文件路径}
        """
        os.makedirs(output_dir, exist_ok=True)
        files = {}

        # 1. TAS 图解数据: SiO2, Na2O+K2O, 分类标签
        tas_data = []
        for s in samples:
            sio2 = s.get("SiO2")
            na2o = s.get("Na2O")
            k2o = s.get("K2O")
            if sio2 and na2o and k2o:
                tas_data.append({
                    "Sample": s.get("Sample.No", ""),
                    "SiO2": sio2,
                    "Na2O+K2O": na2o + k2o,
                    "Type": s.get("Granite type", ""),
                })

        if tas_data:
            tas_path = os.path.join(output_dir, "TAS_diagram_data.csv")
            pd.DataFrame(tas_data).to_csv(tas_path, index=False, encoding='utf-8-sig')
            files["TAS"] = tas_path

        # 2. REE 球粒陨石标准化数据
        ree_fields = ["La", "Ce", "Pr", "Nd", "Sm", "Eu", "Gd", "Tb", "Dy",
                       "Ho", "Er", "Tm", "Yb", "Lu"]
        # Sun & McDonough (1989) 球粒陨石值
        chondrite = {
            "La": 0.237, "Ce": 0.613, "Pr": 0.0928, "Nd": 0.457,
            "Sm": 0.148, "Eu": 0.0563, "Gd": 0.199, "Tb": 0.0361,
            "Dy": 0.246, "Ho": 0.0546, "Er": 0.160, "Tm": 0.0247,
            "Yb": 0.161, "Lu": 0.0246,
        }

        ree_normalized = []
        for s in samples:
            row = {"Sample": s.get("Sample.No", ""), "Type": s.get("Granite type", "")}
            has_data = False
            for f in ree_fields:
                val = s.get(f)
                if val is not None and val > 0:
                    row[f"{f}_norm"] = val / chondrite[f]
                    has_data = True
            if has_data:
                ree_normalized.append(row)

        if ree_normalized:
            ree_path = os.path.join(output_dir, "REE_chondrite_normalized.csv")
            pd.DataFrame(ree_normalized).to_csv(ree_path, index=False, encoding='utf-8-sig')
            files["REE"] = ree_path

        # 3. 花岗岩判别图数据
        disc_data = []
        for s in samples:
            disc_data.append({
                "Sample": s.get("Sample.No", ""),
                "SiO2": s.get("SiO2"),
                "FeOt": s.get("FeOt/(FeOt+MgO)"),
                "ASI": s.get("ASI"),
                "Ga/Al×10⁴": s.get("Ga/Al×10⁴"),
                "Type": s.get("Granite type", ""),
            })

        if disc_data:
            disc_path = os.path.join(output_dir, "discrimination_diagram_data.csv")
            pd.DataFrame(disc_data).to_csv(disc_path, index=False, encoding='utf-8-sig')
            files["discrimination"] = disc_path

        return files


def export_samples_to_excel(samples: list, output_path: str,
                            group_by: str = None) -> str:
    """快捷函数：导出样本到Excel。"""
    engine = ExportEngine()
    if hasattr(samples[0], 'to_flat_dict') if samples else False:
        samples = [s.to_flat_dict() for s in samples]
    return engine.export_excel(samples, output_path, group_by=group_by)
