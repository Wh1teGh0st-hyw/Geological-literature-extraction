"""将现有 Excel 数据迁移到 .gce 项目文件。

支持从现有的 松潘甘孜花岗岩统计.xlsx 等文件导入数据。
"""

import sys
import os
from pathlib import Path
from typing import Optional
from loguru import logger

import pandas as pd
import openpyxl

# 确保能从脚本直接运行
sys.path.insert(0, str(Path(__file__).parent.parent))

from data.database import DatabaseManager, create_database
from data.repository import SampleRepository
from data.models import GeochemSample, MajorElements, TraceElements, REEElements, Isotopes


# ── 列名映射：Excel 列名 → 模型字段路径 ────────────────
# 格式: (excel_column_name, model_field_path)
# model_field_path: "major.sio2" 表示 sample.major.sio2
#                    "trace.li"  表示 sample.trace.li
#                    "mining_area" 表示 sample.mining_area

COLUMN_MAPPING = [
    # 身份信息
    ("矿区", "mining_area"),
    ("岩体", "rock_body"),
    ("Data souce", "data_source"),
    ("Data source", "data_source"),
    ("Sample.No", "sample_no"),
    ("Giranite type", "granite_type"),
    ("Granite type", "granite_type"),
    ("Rock type", "rock_type"),
    # 年代学
    ("成岩年龄", "formation_age"),
    ("error", "age_error"),
    ("Method", "age_method"),
    ("成矿年龄", "mineralization_age"),
    # 空间坐标
    ("X", "longitude"),
    ("Y", "latitude"),
    ("Hight", "elevation"),
    # 主量元素
    ("SiO2", "major.sio2"), ("TiO2", "major.tio2"), ("Al2O3", "major.al2o3"),
    ("Fe2O3", "major.fe2o3"), ("FeO", "major.feo"), ("MnO", "major.mno"),
    ("MgO", "major.mgo"), ("CaO", "major.cao"), ("Na2O", "major.na2o"),
    ("K2O", "major.k2o"), ("P2O5", "major.p2o5"), ("L.O.I", "major.loi"),
    ("Total", "major.total"),
    # 微量元素
    ("Li", "trace.li"), ("Be", "trace.be"), ("Sc", "trace.sc"),
    ("V", "trace.v"), ("Cr", "trace.cr"), ("Co", "trace.co"),
    ("Ni", "trace.ni"), ("Cu", "trace.cu"), ("Zn", "trace.zn"),
    ("Ga", "trace.ga"), ("Rb", "trace.rb"), ("Sr", "trace.sr"),
    ("Y", "trace.y"), ("Zr", "trace.zr"), ("Nb", "trace.nb"),
    ("Cs", "trace.cs"), ("Ba", "trace.ba"), ("Hf", "trace.hf"),
    ("Ta", "trace.ta"), ("Pb", "trace.pb"), ("Th", "trace.th"),
    ("U", "trace.u"),
    # REE
    ("La", "ree.la"), ("Ce", "ree.ce"), ("Pr", "ree.pr"),
    ("Nd", "ree.nd"), ("Sm", "ree.sm"), ("Eu", "ree.eu"),
    ("Gd", "ree.gd"), ("Tb", "ree.tb"), ("Dy", "ree.dy"),
    ("Ho", "ree.ho"), ("Er", "ree.er"), ("Tm", "ree.tm"),
    ("Yb", "ree.yb"), ("Lu", "ree.lu"),
    # 同位素
    ("87Sr/86Sr", "isotopes.sr87_sr86_i"),
    ("ISr(t)", "isotopes.sr87_sr86_i"),
    ("143Nd/144Nd", "isotopes.nd143_nd144_i"),
    ("εNd(t)", "isotopes.epsilon_nd_t"),
    ("176Hf/177Hf", "isotopes.hf176_hf177_i"),
    ("εHf(t)", "isotopes.epsilon_hf_t"),
]


def _safe_float(value) -> Optional[float]:
    """安全地将值转换为 float，解析失败返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    try:
        s = str(value).strip()
        if s in ("", "-", "--", "n.d.", "n.a.", "bdl", "BDL", "N/A", "na", "NA"):
            return None
        # 处理 ± 误差 (如 "220±3")
        if "±" in s:
            s = s.split("±")[0].strip()
        return float(s)
    except (ValueError, TypeError):
        return None


def _set_field(sample: GeochemSample, field_path: str, value) -> None:
    """根据字段路径设置 GeochemSample 的属性。

    Args:
        sample: 目标样本对象
        field_path: 如 "major.sio2", "trace.li", "mining_area"
        value: 要设置的值
    """
    parts = field_path.split(".")
    if len(parts) == 1:
        # 顶层字段
        setattr(sample, parts[0], value)
    elif len(parts) == 2:
        # 嵌套字段, e.g. major.sio2
        group_name, field_name = parts
        group = getattr(sample, group_name)
        setattr(group, field_name, value)
    else:
        logger.warning(f"无法处理的字段路径: {field_path}")


def import_excel_to_db(
    excel_path: str,
    db: DatabaseManager,
    sheet_name: str = "SPGZ ALL",
    header_row: int = 0,
    limit: Optional[int] = None,
) -> int:
    """将 Excel 文件中的数据导入到 SQLite 数据库。

    Args:
        excel_path: Excel 文件路径
        db: 数据库管理器
        sheet_name: 要导入的Sheet名称
        header_row: 表头行号（0-based）
        limit: 限制导入行数（用于测试），None 表示全部
    Returns:
        成功导入的样本数量
    """
    repo = SampleRepository(db)

    # 使用 openpyxl 读取（绕过 pandas 的编码问题）
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet '{sheet_name}' 不存在。可用: {wb.sheetnames}")
        wb.close()
        return 0

    ws = wb[sheet_name]
    logger.info(f"正在读取 Sheet '{sheet_name}': {ws.max_row} 行 × {ws.max_column} 列")

    # 读取表头（第 header_row + 1 行）
    headers = []
    for cell in ws[header_row + 1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # 建立列索引映射: column_name → column_index
    col_index_map = {}
    for idx, h in enumerate(headers):
        col_index_map[h] = idx
        # 同时也映射无空格版本（防止OCR/手工输入差异）
        col_index_map[h.replace(" ", "")] = idx

    # 构建映射: (excel_col_name, model_field_path) → excel_column_index
    mapping_rules = []
    for excel_name, field_path in COLUMN_MAPPING:
        if excel_name in col_index_map:
            mapping_rules.append((col_index_map[excel_name], field_path))

    logger.info(f"找到 {len(mapping_rules)} 个有效列映射")

    # 逐行读取并构建样本
    samples = []
    skipped = 0
    for row_idx in range(header_row + 2, ws.max_row + 1):
        if limit and len(samples) >= limit:
            break

        sample = GeochemSample()
        has_data = False

        for col_idx, field_path in mapping_rules:
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if cell_value is None:
                continue

            # 对于文本字段，保持原值
            if field_path in ("mining_area", "rock_body", "data_source",
                              "sample_no", "granite_type", "rock_type",
                              "age_method", "min_age_method"):
                val = str(cell_value).strip()
                if val and val not in ("None", "none", "-", "--"):
                    _set_field(sample, field_path, val)
                    has_data = True
            else:
                # 数值字段
                val = _safe_float(cell_value)
                if val is not None:
                    _set_field(sample, field_path, val)
                    has_data = True

        # 至少需要有 sample_no 或 sio2 才视为有效数据行
        if sample.sample_no or sample.major.sio2:
            has_data = True

        if has_data:
            samples.append(sample)
            if len(samples) % 200 == 0:
                logger.info(f"  已解析 {len(samples)} 行...")
        else:
            skipped += 1

    wb.close()

    if not samples:
        logger.warning("没有找到有效数据行。请检查列名映射。")
        return 0

    # 批量插入
    count = repo.bulk_insert_samples(samples)
    logger.info(f"导入完成: {count} 条样本, 跳过 {skipped} 个空行")

    # 更新元数据
    db._set_meta("import_source", os.path.basename(excel_path))
    db._set_meta("import_sheet", sheet_name)
    db._set_meta("import_count", str(count))

    return count


def quick_import(excel_path: str, output_path: str, sheet_name: str = "SPGZ ALL") -> int:
    """快捷函数：从 Excel 导入到新的 .gce 文件。

    Args:
        excel_path: 源 Excel 文件路径
        output_path: 目标 .gce 文件路径
        sheet_name: Sheet 名称
    Returns:
        导入样本数
    """
    db = create_database(output_path)
    count = import_excel_to_db(
        excel_path=excel_path,
        db=db,
        sheet_name=sheet_name,
    )
    db.close()
    return count


# ── 命令行入口 ──────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="将现有 Excel 数据迁移到 .gce 项目文件")
    parser.add_argument("excel", help="源 Excel 文件路径")
    parser.add_argument("-o", "--output", default=None, help="目标 .gce 文件路径（默认同名 .gce）")
    parser.add_argument("-s", "--sheet", default="SPGZ ALL", help="Sheet 名称")
    parser.add_argument("-l", "--limit", type=int, default=None, help="限制导入行数（测试用）")
    args = parser.parse_args()

    if args.output is None:
        base = os.path.splitext(args.excel)[0]
        args.output = base + ".gce"

    logger.info(f"源文件: {args.excel}")
    logger.info(f"Sheet: {args.sheet}")
    logger.info(f"目标: {args.output}")

    count = quick_import(args.excel, args.output, args.sheet)
    logger.info(f"迁移完成: {count} 条样本 → {args.output}")
